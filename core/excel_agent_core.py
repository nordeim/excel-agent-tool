#!/usr/bin/env python3
"""
Excel Agent Core Library
Production-grade Excel manipulation with formula validation and template preservation

This is the foundational library used by all CLI tools.
Designed for stateless, security-hardened Excel operations.

Author: Excel Agent Team
License: MIT
Version: 2.0.0
"""

import re
import sys
import json
import subprocess
import tempfile
import shutil
import traceback
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple, Set
from enum import Enum
from datetime import datetime
import threading
import time

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, Protection, NamedStyle
    )
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    from openpyxl.utils import column_index_from_string
    from openpyxl.comments import Comment
except ImportError:
    raise ImportError(
        "openpyxl is required. Install with:\n"
        "  pip install openpyxl\n"
        "  or: uv pip install openpyxl"
    )

# Optional pandas support
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None


# ============================================================================
# EXCEPTIONS
# ============================================================================

class ExcelAgentError(Exception):
    """Base exception for all Excel agent errors."""
    def __init__(self, message: str, details: Optional[Dict[str, Any]] = None):
        super().__init__(message)
        self.message = message
        self.details = details or {}
    
    def to_json(self) -> Dict[str, Any]:
        """Convert exception to JSON-serializable dict."""
        return {
            "error": self.__class__.__name__,
            "message": self.message,
            "details": self.details
        }


class FormulaError(ExcelAgentError):
    """Raised when formula creation, validation, or parsing fails."""
    pass


class InvalidCellReferenceError(ExcelAgentError):
    """Raised when cell reference format is invalid or out of bounds."""
    pass


class ValidationError(ExcelAgentError):
    """Raised when workbook validation fails or produces errors."""
    pass


class SecurityError(ExcelAgentError):
    """Raised when potentially dangerous operations are detected."""
    pass


class FileLockError(ExcelAgentError):
    """Raised when file cannot be locked for exclusive access."""
    pass


# ============================================================================
# CONSTANTS & ENUMS
# ============================================================================

class FormulaErrors(Enum):
    """All possible Excel formula error types."""
    DIV0 = "#DIV/0!"
    REF = "#REF!"
    VALUE = "#VALUE!"
    NAME = "#NAME?"
    NULL = "#NULL!"
    NUM = "#NUM!"
    NA = "#N/A"


# Industry-standard color conventions (RGB hex)
COLOR_INPUT = "0000FF"      # Blue - Hardcoded inputs
COLOR_FORMULA = "000000"    # Black - ALL formulas
COLOR_LINK = "008000"       # Green - Internal workbook links
COLOR_EXTERNAL = "FF0000"   # Red - External file links
COLOR_ASSUMPTION = "FFFF00" # Yellow background - Key assumptions

# Number formatting standards
FORMAT_CURRENCY = '$#,##0_);[Red]($#,##0)'
FORMAT_CURRENCY_DECIMALS = '$#,##0.00_);[Red]($#,##0.00)'
FORMAT_CURRENCY_MM = '$#,##0.0,,_);[Red]($#,##0.0,,)'
FORMAT_PERCENT = "0.0%"
FORMAT_PERCENT_INT = "0%"
FORMAT_YEAR = "@"
FORMAT_MULTIPLE = "0.0x"
FORMAT_NUMBER = "#,##0"
FORMAT_NUMBER_DECIMALS = "#,##0.00"
FORMAT_DATE = "mm/dd/yyyy"
FORMAT_ACCOUNTING = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

# Style names
STYLE_INPUT = "FinancialInput"
STYLE_FORMULA = "FinancialFormula"
STYLE_ASSUMPTION = "FinancialAssumption"

# Validation constants
MAX_FORMULA_LENGTH = 8000
MAX_FORMULA_NESTING = 64
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLS = 16_384


# ============================================================================
# FILE LOCKING
# ============================================================================

class FileLock:
    """Simple file locking mechanism for concurrent access prevention."""
    
    def __init__(self, filepath: Path, timeout: float = 10.0):
        self.filepath = filepath
        self.lockfile = filepath.parent / f".{filepath.name}.lock"
        self.timeout = timeout
        self.acquired = False
    
    def acquire(self) -> bool:
        """Acquire lock with timeout."""
        start_time = time.time()
        while time.time() - start_time < self.timeout:
            try:
                # Try to create lock file exclusively
                self.lockfile.touch(exist_ok=False)
                self.acquired = True
                return True
            except FileExistsError:
                # Lock exists, wait and retry
                time.sleep(0.1)
        
        return False
    
    def release(self) -> None:
        """Release lock."""
        if self.acquired:
            try:
                self.lockfile.unlink(missing_ok=True)
                self.acquired = False
            except Exception:
                pass
    
    def __enter__(self):
        if not self.acquire():
            raise FileLockError(
                f"Could not acquire lock on {self.filepath} within {self.timeout}s"
            )
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.release()
        return False


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def is_valid_cell_reference(ref: str) -> bool:
    """Validates Excel cell reference format (e.g., 'A1', 'BZ5000')."""
    if not ref or not isinstance(ref, str):
        return False
    pattern = r'^[A-Z]{1,3}\d{1,7}$'
    return bool(re.match(pattern, ref.upper()))


def is_valid_range_reference(range_ref: str) -> bool:
    """Validates Excel range reference format (e.g., 'A1:B10')."""
    if not range_ref:
        return False
    
    # Handle sheet-qualified references
    if "!" in range_ref:
        parts = range_ref.split("!")
        if len(parts) != 2:
            return False
        range_ref = parts[1]
    
    if ":" not in range_ref:
        return is_valid_cell_reference(range_ref)
    
    parts = range_ref.split(":")
    if len(parts) != 2:
        return False
    
    return is_valid_cell_reference(parts[0]) and is_valid_cell_reference(parts[1])


def get_cell_coordinates(cell_ref: str) -> Tuple[int, int]:
    """Convert Excel cell reference to (row, column) tuple (1-indexed)."""
    if not is_valid_cell_reference(cell_ref):
        raise InvalidCellReferenceError(f"Invalid cell reference: {cell_ref}")
    
    match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
    if not match:
        raise InvalidCellReferenceError(f"Cannot parse reference: {cell_ref}")
    
    col_str, row_str = match.groups()
    col_num = column_index_from_string(col_str)
    row_num = int(row_str)
    
    if row_num > EXCEL_MAX_ROWS or col_num > EXCEL_MAX_COLS:
        raise InvalidCellReferenceError(f"Reference out of bounds: {cell_ref}")
    
    return row_num, col_num


def get_column_letter(col_num: int) -> str:
    """Convert column number to Excel column letter (1-indexed)."""
    if not 1 <= col_num <= EXCEL_MAX_COLS:
        raise ValueError(f"Column number {col_num} out of range (1-{EXCEL_MAX_COLS})")
    return openpyxl_get_column_letter(col_num)


def parse_range(range_ref: str) -> Tuple[str, str]:
    """Parse range reference into start and end cells."""
    if "!" in range_ref:
        _, range_ref = range_ref.split("!", 1)
    
    if ":" not in range_ref:
        return range_ref, range_ref
    
    start, end = range_ref.split(":", 1)
    return start, end


def is_valid_sheet_name(name: str) -> bool:
    """Validate Excel sheet name."""
    if not name or len(name) > 31:
        return False
    invalid_chars = [':', '\\', '/', '?', '*', '[', ']']
    return not any(char in name for char in invalid_chars)


def sanitize_sheet_name(name: str) -> str:
    """Sanitize sheet name by removing invalid characters."""
    if not name:
        return "Sheet1"
    
    # Remove invalid characters
    for char in [':', '\\', '/', '?', '*', '[', ']']:
        name = name.replace(char, '_')
    
    # Truncate to 31 chars
    name = name[:31]
    
    return name or "Sheet1"


# ============================================================================
# FORMULA SECURITY & VALIDATION
# ============================================================================

def sanitize_formula(formula: str, allow_external: bool = False) -> Tuple[str, List[str]]:
    """
    Sanitize formula for security issues.
    
    Returns:
        Tuple of (sanitized_formula, list_of_warnings)
    """
    warnings = []
    
    # Ensure formula starts with =
    if not formula.startswith('='):
        formula = '=' + formula
    
    # Check for dangerous functions
    dangerous_patterns = [
        (r'WEBSERVICE\s*\(', 'WEBSERVICE function (network access)'),
        (r'HYPERLINK\s*\(', 'HYPERLINK function (potential phishing)'),
        (r'CALL\s*\(', 'CALL function (external DLL execution)'),
        (r'\[[\w\s]+\.xl', 'External workbook reference'),
        (r'INDIRECT\s*\(.*HYPERLINK', 'INDIRECT+HYPERLINK combination (security risk)'),
    ]
    
    for pattern, warning in dangerous_patterns:
        if re.search(pattern, formula, re.IGNORECASE):
            warnings.append(warning)
    
    # Check formula length
    if len(formula) > MAX_FORMULA_LENGTH:
        warnings.append(f'Formula exceeds recommended length ({len(formula)} chars)')
    
    # Check nesting depth
    nesting = formula.count('(') - formula.count(')')
    if abs(nesting) > MAX_FORMULA_NESTING:
        warnings.append(f'Formula nesting depth suspicious ({nesting})')
    
    return formula, warnings


def validate_formula_references(formula: str, existing_sheets: List[str]) -> Tuple[bool, Optional[str]]:
    """Validate all sheet references in formula exist."""
    if not formula or not isinstance(formula, str):
        return False, "Formula must be a non-empty string"
    
    # Extract sheet names from references (e.g., 'Sheet1'!A1 or Sheet1!A1)
    sheet_refs = re.findall(r"'([^']+)'!", formula)
    sheet_refs.extend(re.findall(r"([A-Za-z0-9_]+)!", formula))
    sheet_refs = list(set(sheet_refs))  # Remove duplicates
    
    for sheet_ref in sheet_refs:
        if sheet_ref not in existing_sheets:
            return False, f"Referenced sheet '{sheet_ref}' does not exist"
    
    return True, None


# ============================================================================
# STYLE MANAGEMENT
# ============================================================================

def create_financial_styles(wb: OpenpyxlWorkbook) -> Dict[str, NamedStyle]:
    """Create standard financial modeling styles and register in workbook."""
    styles = {}
    
    # Input style - Blue text
    if STYLE_INPUT not in wb.named_styles:
        input_style = NamedStyle(name=STYLE_INPUT)
        input_style.font = Font(color=COLOR_INPUT)
        input_style.alignment = Alignment(horizontal="left")
        wb.add_named_style(input_style)
        styles[STYLE_INPUT] = input_style
    
    # Formula style - Black text
    if STYLE_FORMULA not in wb.named_styles:
        formula_style = NamedStyle(name=STYLE_FORMULA)
        formula_style.font = Font(color=COLOR_FORMULA)
        formula_style.alignment = Alignment(horizontal="right")
        wb.add_named_style(formula_style)
        styles[STYLE_FORMULA] = formula_style
    
    # Assumption style - Yellow background
    if STYLE_ASSUMPTION not in wb.named_styles:
        assumption_style = NamedStyle(name=STYLE_ASSUMPTION)
        assumption_style.fill = PatternFill(
            start_color=COLOR_ASSUMPTION,
            end_color=COLOR_ASSUMPTION,
            fill_type='solid'
        )
        assumption_style.font = Font(color='000000', bold=True)
        assumption_style.alignment = Alignment(horizontal="center")
        wb.add_named_style(assumption_style)
        styles[STYLE_ASSUMPTION] = assumption_style
    
    return styles


def get_number_format(format_type: str, decimals: int = 2) -> str:
    """Get standard number format string by type."""
    format_map = {
        "currency": FORMAT_CURRENCY_DECIMALS if decimals > 0 else FORMAT_CURRENCY,
        "currency_mm": FORMAT_CURRENCY_MM,
        "percent": FORMAT_PERCENT if decimals > 0 else FORMAT_PERCENT_INT,
        "multiple": FORMAT_MULTIPLE,
        "year": FORMAT_YEAR,
        "number": FORMAT_NUMBER_DECIMALS if decimals > 0 else FORMAT_NUMBER,
        "accounting": FORMAT_ACCOUNTING,
        "date": FORMAT_DATE,
    }
    
    if format_type not in format_map:
        raise ValueError(
            f"Unknown format_type: {format_type}. "
            f"Available: {list(format_map.keys())}"
        )
    
    return format_map[format_type]


# ============================================================================
# VALIDATION ENGINE
# ============================================================================

class ValidationReport:
    """Structured validation report."""
    
    def __init__(
        self,
        status: str,
        total_errors: int = 0,
        total_formulas: int = 0,
        error_summary: Optional[Dict[str, Any]] = None,
        validation_method: str = "unknown"
    ):
        self.status = status
        self.total_errors = total_errors
        self.total_formulas = total_formulas
        self.error_summary = error_summary or {}
        self.validation_method = validation_method
    
    @classmethod
    def success(cls, formulas: int = 0, method: str = "unknown") -> 'ValidationReport':
        """Create success report."""
        return cls('success', total_formulas=formulas, validation_method=method)
    
    @classmethod
    def from_dict(cls, data: dict) -> 'ValidationReport':
        """Create report from dictionary."""
        return cls(
            status=data.get('status', 'unknown'),
            total_errors=data.get('total_errors', 0),
            total_formulas=data.get('total_formulas', 0),
            error_summary=data.get('error_summary', {}),
            validation_method=data.get('validation_method', 'unknown')
        )
    
    def has_errors(self) -> bool:
        """Check if report contains errors."""
        return self.total_errors > 0
    
    def get_error_locations(self, error_type: Optional[str] = None) -> List[str]:
        """Get list of cell locations with errors."""
        locations = []
        for err_type, details in self.error_summary.items():
            if error_type is None or err_type == error_type:
                if isinstance(details, dict):
                    locations.extend(details.get('locations', []))
        return locations
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "status": self.status,
            "total_errors": self.total_errors,
            "total_formulas": self.total_formulas,
            "error_summary": self.error_summary,
            "validation_method": self.validation_method
        }
    
    def __str__(self) -> str:
        if self.status == 'success':
            return f"✅ Validation passed ({self.total_formulas} formulas via {self.validation_method})"
        
        errors = []
        for err_type, details in self.error_summary.items():
            if isinstance(details, dict) and 'count' in details:
                count = details['count']
                errors.append(f"{err_type}: {count}")
        
        return f"❌ Validation failed ({self.validation_method}) - {', '.join(errors)}"


def validate_workbook_python(filepath: Path) -> ValidationReport:
    """
    Pure-Python validator - checks cached error values only.
    Cannot recalculate formulas.
    """
    try:
        wb = load_workbook(filepath, data_only=False)
        
        error_summary = {}
        total_formulas = 0
        
        ERROR_VALUES = {e.value for e in FormulaErrors}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':
                        total_formulas += 1
                        
                        # Check cached value for errors
                        cell_value = str(cell.value) if cell.value else ""
                        
                        if cell_value in ERROR_VALUES:
                            error_type = cell_value
                            if error_type not in error_summary:
                                error_summary[error_type] = {'count': 0, 'locations': []}
                            error_summary[error_type]['count'] += 1
                            error_summary[error_type]['locations'].append(
                                f"{sheet_name}!{cell.coordinate}"
                            )
        
        wb.close()
        
        if error_summary:
            return ValidationReport(
                status='errors_found',
                total_errors=sum(e['count'] for e in error_summary.values()),
                total_formulas=total_formulas,
                error_summary=error_summary,
                validation_method='python_fallback'
            )
        
        return ValidationReport.success(formulas=total_formulas, method='python_fallback')
        
    except Exception as e:
        return ValidationReport(
            status='error',
            error_summary={'error': {'message': str(e)}},
            validation_method='python_fallback'
        )


def check_libreoffice_available() -> bool:
    """Check if LibreOffice is available."""
    try:
        result = subprocess.run(
            ['soffice', '--version'],
            capture_output=True,
            timeout=5
        )
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def validate_workbook(
    filepath: Path,
    method: str = "auto",
    timeout: int = 30
) -> ValidationReport:
    """
    Validate workbook formulas.
    
    Args:
        filepath: Path to Excel file
        method: 'auto', 'libreoffice', or 'python'
        timeout: Timeout in seconds for LibreOffice
    
    Returns:
        ValidationReport
    """
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")
    
    # Auto-detect best method
    if method == "auto":
        if check_libreoffice_available():
            method = "libreoffice"
        else:
            method = "python"
    
    if method == "libreoffice":
        # Note: Full LibreOffice validation would require separate script
        # For now, fall back to Python
        return validate_workbook_python(filepath)
    elif method == "python":
        return validate_workbook_python(filepath)
    else:
        raise ValueError(f"Unknown validation method: {method}")


# ============================================================================
# ERROR REPAIR
# ============================================================================

def repair_errors(
    filepath: Path,
    error_types: Optional[List[str]] = None,
    backup: bool = True
) -> Dict[str, Any]:
    """
    Attempt to repair formula errors.
    
    Args:
        filepath: Excel file path
        error_types: List of error types to repair (None = all)
        backup: Create backup before repair
    
    Returns:
        Dict with repair results
    """
    if backup:
        backup_path = filepath.parent / f"{filepath.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{filepath.suffix}"
        shutil.copy2(filepath, backup_path)
        backup_file = str(backup_path)
    else:
        backup_file = None
    
    results = {
        "repairs_attempted": 0,
        "repairs_successful": 0,
        "backup_file": backup_file,
        "details": {}
    }
    
    try:
        wb = load_workbook(filepath)
        
        # Repair DIV/0 errors
        if error_types is None or "#DIV/0!" in error_types:
            div0_count = 0
            div0_success = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f' and isinstance(cell.value, str):
                            formula = cell.value
                            if 'IFERROR' not in formula.upper():
                                # Wrap in IFERROR
                                if formula.startswith('='):
                                    formula = formula[1:]
                                cell.value = f'=IFERROR({formula}, 0)'
                                div0_count += 1
                                div0_success += 1
            
            results["details"]["#DIV/0!"] = {
                "attempted": div0_count,
                "successful": div0_success,
                "method": "IFERROR wrapper"
            }
            results["repairs_attempted"] += div0_count
            results["repairs_successful"] += div0_success
        
        wb.save(filepath)
        wb.close()
        
    except Exception as e:
        results["error"] = str(e)
    
    return results


# ============================================================================
# MAIN EXCEL AGENT CLASS
# ============================================================================

class ExcelAgent:
    """
    Core Excel manipulation class for stateless tool operations.
    """
    
    def __init__(self, filepath: Optional[Path] = None):
        """Initialize agent with optional file."""
        self.filepath = Path(filepath) if filepath else None
        self.wb: Optional[OpenpyxlWorkbook] = None
        self._lock: Optional[FileLock] = None
    
    def create_new(self, sheets: List[str]) -> None:
        """Create new workbook with specified sheets."""
        self.wb = Workbook()
        # Remove default sheet
        if self.wb.active:
            self.wb.remove(self.wb.active)
        
        # Add requested sheets
        for sheet_name in sheets:
            if not is_valid_sheet_name(sheet_name):
                raise ValueError(f"Invalid sheet name: {sheet_name}")
            self.wb.create_sheet(sheet_name)
        
        # Create financial styles
        create_financial_styles(self.wb)
    
    def open(self, filepath: Path, acquire_lock: bool = True) -> None:
        """Open existing workbook."""
        if not filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
        
        self.filepath = filepath
        
        # Acquire lock if requested
        if acquire_lock:
            self._lock = FileLock(filepath)
            if not self._lock.acquire():
                raise FileLockError(f"Could not lock file: {filepath}")
        
        self.wb = load_workbook(filepath)
        
        # Ensure financial styles exist
        create_financial_styles(self.wb)
    
    def save(self, filepath: Optional[Path] = None) -> None:
        """Save workbook."""
        if not self.wb:
            raise ExcelAgentError("No workbook loaded")
        
        target = filepath or self.filepath
        if not target:
            raise ExcelAgentError("No output path specified")
        
        target = Path(target)
        target.parent.mkdir(parents=True, exist_ok=True)
        
        self.wb.save(target)
        self.filepath = target
    
    def close(self) -> None:
        """Close workbook and release lock."""
        if self.wb:
            try:
                self.wb.close()
            except Exception:
                pass
            self.wb = None
        
        if self._lock:
            self._lock.release()
            self._lock = None
    
    def get_sheet(self, name: str) -> Worksheet:
        """Get worksheet by name."""
        if not self.wb:
            raise ExcelAgentError("No workbook loaded")
        
        if name not in self.wb.sheetnames:
            raise KeyError(f"Sheet '{name}' not found. Available: {self.wb.sheetnames}")
        
        return self.wb[name]
    
    def add_sheet(self, name: str, index: Optional[int] = None) -> Worksheet:
        """Add new worksheet."""
        if not self.wb:
            raise ExcelAgentError("No workbook loaded")
        
        if not is_valid_sheet_name(name):
            raise ValueError(f"Invalid sheet name: {name}")
        
        if name in self.wb.sheetnames:
            raise ValueError(f"Sheet '{name}' already exists")
        
        return self.wb.create_sheet(name, index)
    
    def set_cell_value(
        self,
        sheet: str,
        cell: str,
        value: Any,
        style: Optional[str] = None,
        number_format: Optional[str] = None
    ) -> None:
        """Set cell value with optional style and format."""
        ws = self.get_sheet(sheet)
        target_cell = ws[cell]
        target_cell.value = value
        
        if style and style in self.wb.named_styles:
            target_cell.style = style
        
        if number_format:
            target_cell.number_format = number_format
    
    def add_formula(
        self,
        sheet: str,
        cell: str,
        formula: str,
        validate_refs: bool = True,
        allow_external: bool = False
    ) -> None:
        """Add validated formula to cell."""
        # Sanitize formula
        formula, warnings = sanitize_formula(formula, allow_external)
        
        if warnings and not allow_external:
            raise SecurityError(
                f"Formula contains potentially unsafe operations: {'; '.join(warnings)}"
            )
        
        # Validate references
        if validate_refs:
            is_valid, error = validate_formula_references(formula, self.wb.sheetnames)
            if not is_valid:
                raise FormulaError(f"Invalid reference: {error}")
        
        self.set_cell_value(sheet, cell, formula, style=STYLE_FORMULA)
    
    def add_financial_input(
        self,
        sheet: str,
        cell: str,
        value: Union[int, float],
        comment: Optional[str] = None,
        number_format: Optional[str] = None
    ) -> None:
        """Add financial input with blue style."""
        self.set_cell_value(sheet, cell, value, style=STYLE_INPUT, number_format=number_format)
        
        if comment:
            ws = self.get_sheet(sheet)
            ws[cell].comment = Comment(comment, "ExcelAgent")
    
    def add_assumption(
        self,
        sheet: str,
        cell: str,
        value: Any,
        description: str,
        number_format: Optional[str] = None
    ) -> None:
        """Add assumption with yellow highlight."""
        self.set_cell_value(sheet, cell, value, style=STYLE_ASSUMPTION, number_format=number_format)
        
        ws = self.get_sheet(sheet)
        ws[cell].comment = Comment(description, "ExcelAgent")
    
    def get_value(self, sheet: str, cell: str) -> Any:
        """Get cell value."""
        ws = self.get_sheet(sheet)
        return ws[cell].value
    
    def get_cell_info(self, sheet: str, cell: str) -> Dict[str, Any]:
        """Get comprehensive cell information."""
        ws = self.get_sheet(sheet)
        target_cell = ws[cell]
        
        return {
            "value": target_cell.value,
            "data_type": target_cell.data_type,
            "number_format": target_cell.number_format,
            "is_formula": target_cell.data_type == 'f',
            "comment": target_cell.comment.text if target_cell.comment else None
        }
    
    def apply_range_formula(
        self,
        sheet: str,
        range_ref: str,
        formula_template: str
    ) -> int:
        """
        Apply formula to range.
        
        Args:
            sheet: Sheet name
            range_ref: Range (e.g., "B2:B10")
            formula_template: Formula with {row} and {col} placeholders
        
        Returns:
            Number of cells modified
        """
        ws = self.get_sheet(sheet)
        start_cell, end_cell = parse_range(range_ref)
        
        start_row, start_col = get_cell_coordinates(start_cell)
        end_row, end_col = get_cell_coordinates(end_cell)
        
        count = 0
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                formula = formula_template.replace("{row}", str(row))
                formula = formula.replace("{col}", get_column_letter(col))
                formula = formula.replace("{cell}", f"{get_column_letter(col)}{row}")
                
                if not formula.startswith('='):
                    formula = '=' + formula
                
                ws.cell(row=row, column=col, value=formula)
                count += 1
        
        return count
    
    def format_range(
        self,
        sheet: str,
        range_ref: str,
        number_format: str
    ) -> int:
        """Apply number format to range."""
        ws = self.get_sheet(sheet)
        start_cell, end_cell = parse_range(range_ref)
        
        start_row, start_col = get_cell_coordinates(start_cell)
        end_row, end_col = get_cell_coordinates(end_cell)
        
        count = 0
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).number_format = number_format
                count += 1
        
        return count
    
    def get_workbook_info(self) -> Dict[str, Any]:
        """Get workbook metadata."""
        if not self.wb:
            raise ExcelAgentError("No workbook loaded")
        
        # Count formulas
        total_formulas = 0
        total_cells = 0
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        total_cells += 1
                        if cell.data_type == 'f':
                            total_formulas += 1
        
        info = {
            "sheets": self.wb.sheetnames,
            "sheet_count": len(self.wb.sheetnames),
            "total_formulas": total_formulas,
            "total_cells_with_data": total_cells,
        }
        
        if self.filepath:
            info["file"] = str(self.filepath)
            if self.filepath.exists():
                stat = self.filepath.stat()
                info["file_size_bytes"] = stat.st_size
                info["modified"] = datetime.fromtimestamp(stat.st_mtime).isoformat()
        
        return info
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
        return False


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def create_workbook_from_structure(
    output: Path,
    structure: Dict[str, Any],
    validate: bool = False
) -> Dict[str, Any]:
    """
    Create workbook from structure definition.
    
    Args:
        output: Output file path
        structure: Structure dictionary with keys:
            - sheets: List of sheet names
            - cells: List of cell definitions
            - inputs: List of input definitions
            - assumptions: List of assumption definitions
        validate: Run validation after creation
    
    Returns:
        Dict with creation results
    """
    with ExcelAgent() as agent:
        # Create sheets
        sheets = structure.get("sheets", ["Sheet1"])
        agent.create_new(sheets)
        
        stats = {
            "sheets_created": len(sheets),
            "formulas_added": 0,
            "inputs_added": 0,
            "assumptions_added": 0,
            "cells_set": 0
        }
        
        # Add cells
        for cell_def in structure.get("cells", []):
            sheet = cell_def["sheet"]
            cell = cell_def["cell"]
            
            if "formula" in cell_def:
                agent.add_formula(
                    sheet, cell, cell_def["formula"],
                    allow_external=cell_def.get("allow_external", False)
                )
                stats["formulas_added"] += 1
            elif "value" in cell_def:
                agent.set_cell_value(
                    sheet, cell, cell_def["value"],
                    style=cell_def.get("style"),
                    number_format=cell_def.get("number_format")
                )
                stats["cells_set"] += 1
        
        # Add inputs
        for input_def in structure.get("inputs", []):
            agent.add_financial_input(
                input_def["sheet"],
                input_def["cell"],
                input_def["value"],
                comment=input_def.get("comment"),
                number_format=input_def.get("number_format")
            )
            stats["inputs_added"] += 1
        
        # Add assumptions
        for assumption_def in structure.get("assumptions", []):
            agent.add_assumption(
                assumption_def["sheet"],
                assumption_def["cell"],
                assumption_def["value"],
                assumption_def["description"],
                number_format=assumption_def.get("number_format")
            )
            stats["assumptions_added"] += 1
        
        # Save
        agent.save(output)
        
        # Validate if requested
        if validate:
            validation_result = validate_workbook(output)
            stats["validation_result"] = validation_result.to_dict()
        
        return stats


def export_sheet_to_csv(
    filepath: Path,
    sheet: str,
    output: Path,
    range_ref: Optional[str] = None
) -> int:
    """Export sheet to CSV."""
    import csv
    
    with ExcelAgent(filepath) as agent:
        agent.open(filepath, acquire_lock=False)
        ws = agent.get_sheet(sheet)
        
        if range_ref:
            start_cell, end_cell = parse_range(range_ref)
            start_row, start_col = get_cell_coordinates(start_cell)
            end_row, end_col = get_cell_coordinates(end_cell)
            rows = ws.iter_rows(min_row=start_row, max_row=end_row,
                               min_col=start_col, max_col=end_col)
        else:
            rows = ws.iter_rows()
        
        with open(output, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            row_count = 0
            for row in rows:
                writer.writerow([cell.value for cell in row])
                row_count += 1
        
        return row_count


# ============================================================================
# MODULE METADATA
# ============================================================================

__version__ = "2.0.0"
__all__ = [
    # Core class
    "ExcelAgent",
    
    # Exceptions
    "ExcelAgentError",
    "FormulaError",
    "InvalidCellReferenceError",
    "ValidationError",
    "SecurityError",
    "FileLockError",
    
    # Validation
    "ValidationReport",
    "validate_workbook",
    "repair_errors",
    
    # Utilities
    "is_valid_cell_reference",
    "is_valid_range_reference",
    "get_cell_coordinates",
    "get_column_letter",
    "is_valid_sheet_name",
    "sanitize_sheet_name",
    "get_number_format",
    
    # Convenience functions
    "create_workbook_from_structure",
    "export_sheet_to_csv",
    
    # Constants
    "FormulaErrors",
    "STYLE_INPUT",
    "STYLE_FORMULA",
    "STYLE_ASSUMPTION",
]
